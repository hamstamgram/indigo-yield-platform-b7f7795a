import json
from datetime import datetime, timedelta
import openpyxl

def excel_date_to_datetime(excel_date):
    if isinstance(excel_date, datetime):
        return excel_date.strftime("%Y-%m-%d")
    try:
        val = float(excel_date)
        dt = datetime(1899, 12, 30) + timedelta(days=val)
        return dt.strftime("%Y-%m-%d")
    except:
        return str(excel_date)

wb = openpyxl.load_workbook("Accounting Yield Funds (3).xlsx", data_only=True)

funds_mapping = {
    "SOL Yield Fund": "sol",
    "XRP Yield Fund": "xrp",
    "ETH Yield Fund": "eth",
    "BTC Yield Fund": "btc"
}

output = {}

for sheet_name, fund_key in funds_mapping.items():
    if sheet_name not in wb.sheetnames:
        continue
    
    ws = wb[sheet_name]
    
    # Simple search for rows
    date_row = None
    aum_before_row = None
    top_up_row = None
    investors_row = None
    
    rows = list(ws.iter_rows(min_row=1, max_row=50, values_only=True))
    for i, r in enumerate(rows):
        first_cell = str(r[0]).strip() if r[0] is not None else ""
        if first_cell == "Investors":
            date_row = r
            investors_row = rows[i+1:] # Might need to guess who investors are
        elif first_cell == "AUM Before":
            aum_before_row = r
        elif first_cell == "Top Up / Withdrawals":
            top_up_row = r

    if not date_row or not aum_before_row or not top_up_row:
        continue
        
    actions = []
    
    for col_idx in range(3, len(date_row)): # D onwards
        val = date_row[col_idx]
        if val is None:
            continue
            
        date_str = excel_date_to_datetime(val)
        aum_before = aum_before_row[col_idx]
        if aum_before is None:
            aum_before = 0
            
        top_up = top_up_row[col_idx]
        if top_up is None:
            top_up = 0
            
        try:
            tu_val = float(top_up)
            has_tx = abs(tu_val) > 0.0001
        except:
            has_tx = False
            
        # Is this an end of month report?
        is_report = False
        if "-28" in date_str or "-29" in date_str or "-30" in date_str or "-31" in date_str or "-01" in date_str:
            if not has_tx:
                is_report = True

        if is_report:
            actions.append({
                "action": "yield",
                "date": date_str,
                "type": "Reporting",
                "aum": aum_before
            })
        else:
            actions.append({
                "action": "yield",
                "date": date_str,
                "type": "Transaction",
                "aum": aum_before
            })
            
            # Now find who did the transaction
            for inv_row in investors_row:
                inv_name = inv_row[0]
                if not inv_name or inv_name in ["Total AUM", "Indigo Fees"]:
                    continue
                    
                cell_val = inv_row[col_idx]
                if cell_val is not None:
                    # Check if cell formula had a manual top up
                    # Wait, data_only=True means we only get values.
                    # We need the real excel dump for the comments!
                    pass

print("Use excel_dump.json instead for accuracy on exact names and comments.")
