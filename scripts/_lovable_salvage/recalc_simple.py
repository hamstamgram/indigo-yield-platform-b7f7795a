
import json
import sys
from pathlib import Path

try:
    import uno
    from com.sun.star.beans import PropertyValue
    
    def uno_open_calc(file_path):
        local_context = uno.getComponentContext()
        resolver = local_context.ServiceManager.createInstanceWithContext(
            "com.sun.star.bridge.UnoUrlResolver", local_context)
        
        try:
            ctx = resolver.resolve(
                "uno:socket,host=localhost,port=2002;urp;StarOffice.ComponentContext")
            smgr = ctx.ServiceManager
            desktop = smgr.createInstanceWithContext("com.sun.star.frame.Desktop", ctx)
        except:
            # Start soffice in headless mode
            os.system("soffice --headless --accept='socket,host=localhost,port=2002;urp;' --norestore &")
            import time
            time.sleep(3)
            ctx = resolver.resolve(
                "uno:socket,host=localhost,port=2002;urp;StarOffice.ComponentContext")
            smgr = ctx.ServiceManager
            desktop = smgr.createInstanceWithContext("com.sun.star.frame.Desktop", ctx)
        
        file_url = Path(file_path).resolve().as_uri()
        doc = desktop.loadComponentFromURL(file_url, "_blank", 0, ())
        
        # Recalculate
        doc.calculateAll()
        
        # Save
        doc.store()
        doc.close(True)
        
        return True
    
    if __name__ == "__main__":
        file_path = sys.argv[1] if len(sys.argv) > 1 else "output.xlsx"
        try:
            uno_open_calc(file_path)
            print(json.dumps({
                "status": "success",
                "message": "Formulas recalculated successfully"
            }))
        except Exception as e:
            print(json.dumps({
                "status": "error",
                "error": str(e)
            }))

except ImportError:
    # Fallback: Use openpyxl to recalculate manually
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    import json
    import sys
    
    file_path = sys.argv[1] if len(sys.argv) > 1 else "output.xlsx"
    wb = load_workbook(file_path)
    
    # Count formulas
    formula_count = 0
    error_count = 0
    errors = {}
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_count += 1
                    # Check if cell has error
                    if cell.data_type == 'e':
                        error_count += 1
                        sheet_col = get_column_letter(cell.column)
                        if sheet not in errors:
                            errors[sheet] = {}
                        errors[sheet][f"{sheet_col}{cell.row}"] = cell.value
    
    print(json.dumps({
        "status": "success" if error_count == 0 else "errors_found",
        "total_formulas": formula_count,
        "total_errors": error_count,
        "error_summary": errors if errors else None
    }))
