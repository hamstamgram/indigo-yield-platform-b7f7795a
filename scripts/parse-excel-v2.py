#!/usr/bin/env python3
"""
Parse indigo-accounting.xlsx into a structured event JSON for seed-excel-v2.ts.
Output: scripts/seed-data/excel-events-v2.json

BTC (rows-first, rows 2-19 only):
- Row 1 cols L+: snapshot dates
- Rows 2-19: event rows (col A=date, col B=AUM, col C=gross_pct, col G=comment, col I=investor, col J=fee, col K=IB)
- Each row = one investor + one event. Multiple rows may share same event date/gross_pct.
- Snapshot columns = investor balance at that date (right after all events on that date)

ETH/USDT/SOL/XRP (columns-first):
- Row 1: AUM Before, Row 2: Top Up, Row 3: AUM After, Row 4: Gross%, Row 7: Comments
- Row 8: header with dates in cols D+
- Rows 9 to 'Total': investor rows
"""

import openpyxl
import json
import sys
import calendar
from datetime import datetime, date
from collections import defaultdict, OrderedDict
from openpyxl.utils import get_column_letter

EXCEL_PATH = '/Users/mama/indigo-yield-platform-v01/scripts/seed-data/indigo-accounting.xlsx'
OUTPUT_PATH = '/Users/mama/indigo-yield-platform-v01/scripts/seed-data/excel-events-v2.json'

INVESTOR_UUID_MAP = {
    'indigo fees': 'b464a3f7-60d5-4bc0-9833-7b413bcc6cae',
    'indigo fees account': 'b464a3f7-60d5-4bc0-9833-7b413bcc6cae',
    'indigo fee': 'b464a3f7-60d5-4bc0-9833-7b413bcc6cae',
    'jose molla': '203caf71-a9ac-4e2a-bbd3-b45dd51758d4',
    'jose': '203caf71-a9ac-4e2a-bbd3-b45dd51758d4',
    'kyle gulamerian': 'b4f5d56b-b128-4799-b805-d34264165f45',
    'kyle': 'b4f5d56b-b128-4799-b805-d34264165f45',
    'matthias reiser': 'd8643c68-7045-458a-b105-a41f56085c55',
    'mathias': 'd8643c68-7045-458a-b105-a41f56085c55',
    'mathias reiser': 'd8643c68-7045-458a-b105-a41f56085c55',
    'thomas puech': '44801beb-4476-4a9b-9751-4e70267f6953',
    'danielle richetta': 'e134e0df-d4e7-49c4-80b3-4ef37af6bebf',
    'family kabbaj': 'f917cd8b-2d12-428c-ae3c-210b7ee3ae75',
    'kabbaj': 'f917cd8b-2d12-428c-ae3c-210b7ee3ae75',
    'kabbaj fam': 'f917cd8b-2d12-428c-ae3c-210b7ee3ae75',
    'victoria pariente-cohen': '249f4ab3-3433-4d81-ac92-1531b3573a50',
    'victoria': '249f4ab3-3433-4d81-ac92-1531b3573a50',
    'nathanael cohen': 'ed91c89d-23de-4981-b6b7-60e13f1a6767',
    'nathanaël cohen': 'ed91c89d-23de-4981-b6b7-60e13f1a6767',
    'blondish': '529cac24-615c-4408-b683-2c4ab635d6fd',
    'oliver loisel': 'fbf8e2f4-7c5d-4496-a486-f0d8e88cc794',
    'paul johnson': 'd1f8c666-58c5-4a83-a5c6-0f66a380aaf2',
    'alex jacobs': 'd681a28c-bb59-4bb7-bf34-ab23910596df',
    'alex\xa0jacobs': 'd681a28c-bb59-4bb7-bf34-ab23910596df',
    'sam johnson': '2f7b8bb2-6a60-4fc9-953d-b9fae44337c1',
    'sam johnson ': '2f7b8bb2-6a60-4fc9-953d-b9fae44337c1',
    'ryan van der wall': 'f462d9e5-7363-4c82-a144-4e694d2b55da',
    'nath & thomas': '99e5a116-44ba-4a45-9f56-5877b235f960',
    'vivie & liana': '981dd85c-35c8-4254-a3e9-27c2af302815',
    'nsvo holdings': '114164b0-1aba-4b40-9abc-8d72adfdc60a',
    'babak eftekhari': 'cdcccf6e-32f9-475a-9f88-34272ca3e64b',
    'lars ahlgreen': '9405071c-0b52-4399-85da-9f1ba9b289c1',
    'indigo digital asset fund lp': 'd91f3eb7-bd47-4c42-ab4f-c4f20fb41b13',
    'indigo main fund': 'd91f3eb7-bd47-4c42-ab4f-c4f20fb41b13',
    'indigo lp': 'd91f3eb7-bd47-4c42-ab4f-c4f20fb41b13',
    'julien grunebaum': '7fdedf56-e838-45ea-91f8-6e441810c761',
    'daniele francilia': 'd1f39136-4d87-4e7f-8885-a413c21d9a56',
    'pierre bezençon': '511991c7-93a2-4d2b-b42a-43120d58f672',
    'pierre bezencon': '511991c7-93a2-4d2b-b42a-43120d58f672',
    'matthew beatty': '24f3054e-a125-4954-8861-55aa617cbb2c',
    'bo de kriek': '98dd4ff5-b5cb-4257-a501-aa25a6d638c5',
    'bo de\xa0kriek': '98dd4ff5-b5cb-4257-a501-aa25a6d638c5',
    'bo kriek': '98dd4ff5-b5cb-4257-a501-aa25a6d638c5',
    'dario deiana': 'bb655a37-9e91-4166-b575-cafbbbb8c200',
    'alain bensimon': '20396ec2-c919-46ef-b3a3-8005a8a34bd3',
    'anne cecile noique': '64cb831a-3365-4a89-9369-620ab7a1ff26',
    'terance chen': '3705c2cd-49d2-4e3b-ac09-7c1f98ebb93c',
    'indigo ventures': 'af16a20d-df70-4357-9b10-efcd25d0c1aa',
    'halley86': '32d75475-0b78-4b7b-925a-e9429f6fe66d',
    'advantage blockchain': '3a9a5615-d4dd-4b7f-ae20-3d48a7de3dcc',
    'alec beckman': '5fc170e2-7a07-4f32-991f-d8b6deec277c',
    'tomer zur': '82f58ac0-2d34-4c00-b0df-34383c1d1dfd',
    'brandon hood': 'a00073d1-f37d-4e21-a54b-1b55df17e85a',
    'sacha oshry': 'd5719d57-5308-4b9d-8a4f-a9a8aa596af4',
    'monica levy chicheportiche': 'c85bddf5-7720-47a5-8336-669ea604b94b',
    'valeria cruz': 'e9bbc28b-5d8d-410c-940b-b37a54a726e0',
    'ventures life style': '7d049f7f-b77f-4650-b772-6a8806f00103',
    'joel barbeau': '99e56523-32a6-43e5-b9b3-789992cc347c',
    'joel\xa0barbeau': '99e56523-32a6-43e5-b9b3-789992cc347c',
}

SKIP_INVESTORS = {
    # These investors are now in the UUID map with real accounts
}

# Investors that receive yield/IB credits passively but don't make real deposits/withdrawals.
# Their balances in Excel reflect accumulated yield+IB commissions, not real manual transactions.
# The V5 yield engine handles IB_CREDIT and FEE_CREDIT for these automatically.
# IB relationships: Babak->Lars, Advantage->Alec, Ventures LifeStyle->Joel, Sam Johnson->Ryan, Paul Johnson->Alex Jacobs
YIELD_ONLY_INVESTORS = {
    'b464a3f7-60d5-4bc0-9833-7b413bcc6cae',  # Indigo Fees (receives FEE_CREDIT)
    '9405071c-0b52-4399-85da-9f1ba9b289c1',  # Lars Ahlgreen (IB for Babak Eftekhari)
    '5fc170e2-7a07-4f32-991f-d8b6deec277c',  # Alec Beckman (IB for Advantage Blockchain)
    '99e56523-32a6-43e5-b9b3-789992cc347c',  # Joel Barbeau (IB for Ventures Life Style)
    'f462d9e5-7363-4c82-a144-4e694d2b55da',  # Ryan Van Der Wall (IB for Sam Johnson)
    'd681a28c-bb59-4bb7-bf34-ab23910596df',  # Alex Jacobs (IB for Paul Johnson)
}

FUND_IDS = {
    'BTC': '0a048d9b-c4cf-46eb-b428-59e10307df93',
    'ETH': '717614a2-9e24-4abc-a89d-02209a3a772a',
    'SOL': '7574bc81-aab3-4175-9e7f-803aa6f9eb8f',
    'USDT': '8ef9dc49-e76c-4882-84ab-a449ef4326db',
    'XRP': '2c123c4f-76b4-4504-867e-059649855417',
}

ADMIN_ID = 'a16a7e50-fefd-4bfe-897c-d16279b457c2'

# Fee metadata per investor per fund
INVESTOR_METADATA = {}  # uuid -> {fee_pct, ib_pct}


def normalize_name(name):
    if name is None:
        return None
    return str(name).strip().lower().replace('\xa0', ' ')


def resolve_uuid(name):
    if name is None:
        return None
    key = normalize_name(name)
    if key in SKIP_INVESTORS:
        return None
    if key in ('total', 'total aum', 'x'):
        return None
    uuid = INVESTOR_UUID_MAP.get(key)
    if uuid is None:
        print(f"  WARNING: No UUID for '{name}' (key='{key}')", file=sys.stderr)
    return uuid


def to_date_str(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d') if isinstance(val, datetime) else val.isoformat()
    return str(val)[:10]


def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def fmt(val):
    """Format number as string for Decimal precision."""
    if val is None:
        return None
    f = safe_float(val)
    if f is None:
        return None
    return f'{f:.10f}'


def last_day(date_str):
    year, month, _ = map(int, date_str.split('-'))
    d = calendar.monthrange(year, month)[1]
    return f'{year:04d}-{month:02d}-{d:02d}'


def is_flow(gp):
    if gp is None:
        return False
    return abs(float(gp) - 1.0) < 0.00001


def is_yield(gp):
    if gp is None:
        return False
    f = float(gp)
    return 0.00001 < f < 0.9999


def is_zero(gp):
    if gp is None:
        return False
    return abs(float(gp)) < 0.00001


def parse_btc(ws):
    """Parse BTC Yield Fund (rows-first layout, rows 2-79)."""
    print("Parsing BTC Yield Fund...")

    # Get snapshot dates from row 1, cols L+ (index 12+)
    snapshot_dates = []
    for col_idx in range(12, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value is not None:
            snapshot_dates.append((col_idx, to_date_str(cell.value)))

    snap_list = [sd for _, sd in snapshot_dates]
    date_to_col = {}
    for col, sd in snapshot_dates:
        if sd not in date_to_col:
            date_to_col[sd] = col

    print(f"  Snapshots: {len(snapshot_dates)} ({snap_list[0]} to {snap_list[-1]})")

    # Read all investor rows (2-19) to get per-investor balance history
    investor_meta = {}   # name -> {uuid, fee_pct, ib_pct}
    investor_balances = {}  # name -> {date -> balance} (for date-based lookups)
    investor_col_balances = {}  # name -> {col_idx -> balance} (for column-based lookups)

    for row_idx in range(2, 20):
        name_raw = ws.cell(row=row_idx, column=9).value
        if not name_raw or str(name_raw) == 'Total AUM':
            continue

        name = str(name_raw)
        uuid = resolve_uuid(name)
        fee_pct = safe_float(ws.cell(row=row_idx, column=10).value)
        ib_pct = safe_float(ws.cell(row=row_idx, column=11).value)

        investor_meta[name] = {'uuid': uuid, 'fee_pct': fee_pct, 'ib_pct': ib_pct}

        if uuid:
            INVESTOR_METADATA[uuid] = {'fee_pct': fee_pct, 'ib_pct': ib_pct}
            if name not in investor_balances:
                investor_balances[name] = {}
                investor_col_balances[name] = {}
            for col, sd in snapshot_dates:
                val = safe_float(ws.cell(row=row_idx, column=col).value)
                val = val if val is not None else 0.0
                investor_balances[name][sd] = val
                investor_col_balances[name][col] = val

    # Build ordered event list from rows 2-79 (by row order = chronological).
    # Rows 2-19:  section 1 investors (have name in col I, full balance history)
    # Row 20:     "Total AUM" totals row — also a valid FLOW event (Thomas TAC withdrawal)
    # Rows 22-39: section 2 metadata rows — may have no date/AUM data (skipped)
    # Rows 40-79: event-only rows — no investor name in col I, amounts from comments
    # Key: (date, gp, aum_rounded) to distinguish same-date same-type events with different AUMs
    events_ordered = OrderedDict()
    for row_idx in range(2, 80):
        ev_date_raw = ws.cell(row=row_idx, column=1).value
        aum_raw = ws.cell(row=row_idx, column=2).value
        gp_raw = ws.cell(row=row_idx, column=3).value
        comment_raw = ws.cell(row=row_idx, column=7).value

        if ev_date_raw is None or aum_raw is None:
            continue

        ev_date = to_date_str(ev_date_raw)
        gp = safe_float(gp_raw)
        aum = safe_float(aum_raw)
        # Round AUM to 4 decimal places to distinguish different events
        aum_key = round(aum, 4) if aum else 0
        key = (ev_date, gp, aum_key)

        if key not in events_ordered:
            events_ordered[key] = {
                'date': ev_date,
                'gross_pct': gp,
                'aum': aum,
                'aum_key': aum_key,
                'comment': str(comment_raw).replace('\n', ' ') if comment_raw else '',
                'row_idx': row_idx,
            }

    print(f"  Distinct events (rows 2-79): {len(events_ordered)}")

    # Associate each event with its correct snapshot column.
    # For dates with multiple snapshot columns, assign flow events the snapshot cols in order.
    # Yield events on same-day-as-flow get a snap_col too, but won't read positions from it.
    date_snap_cols = defaultdict(list)  # date -> [col_idx, ...]
    for col, sd in snapshot_dates:
        date_snap_cols[sd].append(col)

    # Assign snapshot columns to events.
    # ONLY FLOW events consume snap_cols (in row order).
    # YIELD events get the FIRST snap_col for their date (they use it only for non-same-day-flow dates).
    events_list_ordered = list(events_ordered.values())

    # Build per-date flow event list in order
    date_flow_events = defaultdict(list)
    for ev in events_list_ordered:
        if is_flow(ev['gross_pct']):
            date_flow_events[ev['date']].append(ev)

    # Assign snap_cols to flow events in order
    for d, flow_evs in date_flow_events.items():
        snap_cols = date_snap_cols.get(d, [])
        for i, ev in enumerate(flow_evs):
            if i < len(snap_cols):
                ev['snap_col'] = snap_cols[i]
            else:
                ev['snap_col'] = snap_cols[-1] if snap_cols else None

    # Assign snap_cols to YIELD events (first col for their date)
    for ev in events_list_ordered:
        if not is_flow(ev['gross_pct']):
            snap_cols = date_snap_cols.get(ev['date'], [])
            ev['snap_col'] = snap_cols[0] if snap_cols else None

    # Track expected position per investor (after yield adjustments)
    positions = {}  # name -> current position

    # Pre-pass: identify which dates have BOTH yield and flow events (same day)
    dates_with_yield_and_flow = set()
    dates_events_gps = defaultdict(list)
    for ev in events_list_ordered:
        dates_events_gps[ev['date']].append(ev['gross_pct'])
    for d, gps in dates_events_gps.items():
        has_yield = any(is_yield(g) for g in gps)
        has_flow = any(is_flow(g) for g in gps)
        if has_yield and has_flow:
            dates_with_yield_and_flow.add(d)

    # Dates that have ANY flow event (including zero-yield+flow combinations).
    # Used to prevent zero-yield position updates from pre-empting balance-delta detection.
    dates_with_flow = {
        d for d, gps in dates_events_gps.items()
        if any(is_flow(g) for g in gps)
    }

    output_events = []

    for key, ev in events_ordered.items():
        ev_date = ev['date']
        gp = ev['gross_pct']
        snap_col = ev.get('snap_col')

        if is_zero(gp) and gp is not None:
            # Zero yield: update positions from snapshot if available.
            # SKIP position update if this date also has a FLOW event: the FLOW event's
            # balance-delta will correctly detect deposits/withdrawals using the snapshot.
            # Pre-updating positions here would erase the delta and make FLOW look like a no-op.
            if snap_col and ev_date not in dates_with_flow:
                for name in investor_balances:
                    bal = investor_balances[name].get(ev_date, 0.0)
                    if bal > 0:
                        positions[name] = bal
            print(f"  BTC: Skipping zero yield on {ev_date}")
            continue

        if is_flow(gp):
            # FLOW EVENT: find deposits/withdrawals by checking ALL investor snapshots.
            # Key challenge: when YIELD and FLOW both happen on the same date (in that order),
            # the snapshot captures the post-yield AND post-flow balance.

            if snap_col is None and ev_date != '2024-07-01':
                # No snapshot column for this date: can't use balance-delta approach.
                # Parse transactions from the event comment instead.
                flow_transactions = []
                comment = ev.get('comment', '')

                if 'Thomas' in comment and 'TAC' in comment.upper():
                    # Thomas FULL withdrawal to TAC program
                    thomas_uuid = resolve_uuid('Thomas Puech')
                    thomas_meta = investor_meta.get('Thomas Puech', {})
                    if thomas_uuid:
                        flow_transactions.append({
                            'type': 'WITHDRAWAL',
                            'investor_name': 'Thomas Puech',
                            'investor_uuid': thomas_uuid,
                            'amount': 'FULL',
                            'fee_pct': thomas_meta.get('fee_pct'),
                            'ib_pct': thomas_meta.get('ib_pct'),
                        })
                        positions['Thomas Puech'] = 0.0

                if 'Danielle' in comment and '0.13' in comment:
                    # Danielle partial withdrawal of 0.13 BTC
                    danielle_uuid = resolve_uuid('Danielle Richetta')
                    danielle_meta = investor_meta.get('Danielle Richetta', {})
                    if danielle_uuid:
                        flow_transactions.append({
                            'type': 'WITHDRAWAL',
                            'investor_name': 'Danielle Richetta',
                            'investor_uuid': danielle_uuid,
                            'amount': '0.1300000000',
                            'fee_pct': danielle_meta.get('fee_pct'),
                            'ib_pct': danielle_meta.get('ib_pct'),
                        })
                        positions['Danielle Richetta'] = positions.get('Danielle Richetta', 0.0) - 0.13

                if flow_transactions:
                    output_events.append({
                        'fund': 'BTC',
                        'fund_id': FUND_IDS['BTC'],
                        'event_type': 'FLOW',
                        'date': ev_date,
                        'aum_after': fmt(ev['aum']),
                        'comment': ev['comment'],
                        'transactions': flow_transactions,
                    })
                else:
                    print(f"  BTC: FLOW {ev_date} snap_col=None, comment not matched: {comment[:80]}")

            else:
                # Normal balance-delta approach: compare each investor's snapshot balance
                # to their current tracked position to detect deposits/withdrawals.

                # For same-day yield+flow: we need the YIELD gross_pct from the same date
                same_day_yield_gp = None
                if ev_date in dates_with_yield_and_flow:
                    for (d, g, a) in events_ordered.keys():
                        if d == ev_date and is_yield(g):
                            same_day_yield_gp = g
                            break

                flow_transactions = []

                for name, balances in investor_balances.items():
                    meta = investor_meta.get(name, {})
                    uuid = meta.get('uuid')
                    if uuid is None:
                        continue

                    # Skip investors who only receive yield (no real deposits/withdrawals)
                    if uuid in YIELD_ONLY_INVESTORS:
                        if snap_col:
                            col_balances = investor_col_balances.get(name, {})
                            positions[name] = col_balances.get(snap_col, 0.0)
                        continue

                    # Skip initial setup (no snapshot for 2024-07-01)
                    if ev_date == '2024-07-01' and snap_col is None:
                        continue

                    # Use column-index-based lookup to handle duplicate dates correctly
                    col_balances = investor_col_balances.get(name, {})
                    curr = col_balances.get(snap_col, 0.0) if snap_col else 0.0
                    pre_flow_position = positions.get(name, 0.0)

                    # If there was a yield on the same date BEFORE the flow,
                    # the positions[] already reflects post-yield state.
                    # So pre_flow_position already accounts for yield.
                    # We just compare pre_flow_position to curr.

                    if curr == 0.0 and pre_flow_position == 0.0:
                        continue

                    if pre_flow_position < 0.0001 and curr > 0.01:
                        # New deposit (started at 0, now has balance)
                        flow_transactions.append({
                            'type': 'DEPOSIT',
                            'investor_name': name,
                            'investor_uuid': uuid,
                            'amount': fmt(curr),
                            'fee_pct': meta.get('fee_pct'),
                            'ib_pct': meta.get('ib_pct'),
                        })
                        positions[name] = curr
                    elif curr < 0.0001 and pre_flow_position > 0.001:
                        # Full withdrawal (position goes to 0)
                        flow_transactions.append({
                            'type': 'WITHDRAWAL',
                            'investor_name': name,
                            'investor_uuid': uuid,
                            'amount': 'FULL',
                            'fee_pct': meta.get('fee_pct'),
                            'ib_pct': meta.get('ib_pct'),
                        })
                        positions[name] = 0.0
                    elif curr > 0.0001 and pre_flow_position > 0.001:
                        # Both non-zero: compute the flow amount
                        flow_amount = curr - pre_flow_position
                        abs_flow = abs(flow_amount)
                        # Flow threshold: > 0.05 BTC or > 1% of position
                        pct_change = abs_flow / max(pre_flow_position, 0.001)
                        if abs_flow > 0.05 or pct_change > 0.01:
                            if flow_amount < 0:
                                flow_transactions.append({
                                    'type': 'WITHDRAWAL',
                                    'investor_name': name,
                                    'investor_uuid': uuid,
                                    'amount': fmt(abs_flow),
                                    'fee_pct': meta.get('fee_pct'),
                                    'ib_pct': meta.get('ib_pct'),
                                })
                            else:
                                flow_transactions.append({
                                    'type': 'DEPOSIT',
                                    'investor_name': name,
                                    'investor_uuid': uuid,
                                    'amount': fmt(flow_amount),
                                    'fee_pct': meta.get('fee_pct'),
                                    'ib_pct': meta.get('ib_pct'),
                                })
                        positions[name] = curr
                    else:
                        if snap_col:
                            positions[name] = curr

                # Handle special case: 2024-07-01 initial setup
                if ev_date == '2024-07-01':
                    # Jose deposits 3.468 BTC
                    jose_uuid = resolve_uuid('Jose Molla')
                    jose_meta = investor_meta.get('Jose Molla', {})
                    flow_transactions = [{
                        'type': 'DEPOSIT',
                        'investor_name': 'Jose Molla',
                        'investor_uuid': jose_uuid,
                        'amount': fmt(3.468),
                        'fee_pct': jose_meta.get('fee_pct'),
                        'ib_pct': jose_meta.get('ib_pct'),
                    }]
                    # Initialize positions from first available snapshots
                    for name, balances in investor_balances.items():
                        # Use 2024-08-01 as initial reference after first yield
                        bal_aug1 = balances.get('2024-08-01', 0.0)
                        if bal_aug1 > 0:
                            positions[name] = bal_aug1

                if flow_transactions:
                    output_events.append({
                        'fund': 'BTC',
                        'fund_id': FUND_IDS['BTC'],
                        'event_type': 'FLOW',
                        'date': ev_date,
                        'aum_after': fmt(ev['aum']),
                        'comment': ev['comment'],
                        'transactions': flow_transactions,
                    })
                else:
                    print(f"  BTC: FLOW on {ev_date} no transactions (comment: {ev['comment'][:60]})")

        elif is_yield(gp):
            # YIELD EVENT: record for V5 engine
            period_end = last_day(ev_date)

            # Update positions after yield:
            # If same-day flow follows this yield event, the snapshot on this date
            # includes the post-flow state, not just post-yield.
            # In that case, apply yield rate mathematically to avoid picking up the flow.
            if ev_date in dates_with_yield_and_flow:
                # Apply yield rate to each investor's position
                # Net yield for investor = gross_pct * (1 - fee_pct)
                for name in investor_balances:
                    meta = investor_meta.get(name, {})
                    if positions.get(name, 0.0) > 0.0001:
                        fee = meta.get('fee_pct') or 0.3
                        net_yield_rate = gp * (1 - fee)
                        new_pos = positions[name] * (1 + net_yield_rate)
                        positions[name] = new_pos
            else:
                # No same-day flow: use snapshot directly (it's accurate)
                if snap_col:
                    for name in investor_balances:
                        col_bals = investor_col_balances.get(name, {})
                        bal = col_bals.get(snap_col, 0.0)
                        if bal is not None:
                            positions[name] = bal

            output_events.append({
                'fund': 'BTC',
                'fund_id': FUND_IDS['BTC'],
                'event_type': 'YIELD',
                'date': ev_date,
                'period_end': period_end,
                'aum_after': fmt(ev['aum']),
                'gross_pct': gp,
                'comment': ev['comment'],
            })

    print(f"  BTC: {len(output_events)} output events ({sum(1 for e in output_events if e['event_type']=='FLOW')} flow, {sum(1 for e in output_events if e['event_type']=='YIELD')} yield)")
    return output_events


def parse_btc_final_balances(ws):
    """Extract final Excel balances for BTC investors."""
    snapshot_dates = []
    for col_idx in range(12, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value is not None:
            snapshot_dates.append((col_idx, to_date_str(cell.value)))

    if not snapshot_dates:
        return {}

    last_col, last_date = snapshot_dates[-1]
    result = {}

    for row_idx in range(2, 20):
        name_raw = ws.cell(row=row_idx, column=9).value
        if not name_raw or str(name_raw) == 'Total AUM':
            continue
        uuid = resolve_uuid(str(name_raw))
        if uuid:
            bal = safe_float(ws.cell(row=row_idx, column=last_col).value)
            result[uuid] = {
                'name': str(name_raw),
                'balance': bal if bal else 0.0,
                'as_of': last_date,
            }

    return result


def parse_cf(ws, fund_name, fund_id):
    """Parse columns-first sheet (ETH/USDT/SOL/XRP)."""
    print(f"Parsing {fund_name} Yield Fund...")

    # Get event dates from row 8, cols D+
    event_cols = []
    for col_idx in range(4, ws.max_column + 1):
        cell = ws.cell(row=8, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (datetime, date)):
            event_cols.append({'col_idx': col_idx, 'date': to_date_str(cell.value)})

    if not event_cols:
        print(f"  WARNING: No event columns for {fund_name}")
        return []

    print(f"  Event columns: {len(event_cols)} ({event_cols[0]['date']} to {event_cols[-1]['date']})")

    # Read event metadata (rows 1-7)
    events_meta = []
    for ec in event_cols:
        ci = ec['col_idx']
        gp = safe_float(ws.cell(row=4, column=ci).value)
        top_up = safe_float(ws.cell(row=2, column=ci).value)
        aum_after = safe_float(ws.cell(row=3, column=ci).value)
        comment_raw = ws.cell(row=7, column=ci).value
        comment = str(comment_raw).replace('\n', ' ') if comment_raw else ''

        events_meta.append({
            'col_idx': ci,
            'date': ec['date'],
            'gross_pct': gp,
            'top_up': top_up,
            'aum_after': aum_after,
            'comment': comment,
        })

    # Get investors from rows 9 to 'Total' row
    investors = []
    for row_idx in range(9, ws.max_row + 1):
        name_raw = ws.cell(row=row_idx, column=1).value
        if name_raw is None:
            continue
        name_str = str(name_raw).strip()
        name_lower = normalize_name(name_str)

        if 'total' in name_lower:
            break
        if name_str == 'x' or name_lower == 'x':
            continue
        if name_lower in SKIP_INVESTORS:
            continue

        fee_pct = safe_float(ws.cell(row=row_idx, column=2).value)
        ib_raw = ws.cell(row=row_idx, column=3).value
        ib_pct = safe_float(ib_raw) if isinstance(ib_raw, (int, float)) else None

        uuid = resolve_uuid(name_str)

        balances = {}
        for ec in event_cols:
            val = safe_float(ws.cell(row=row_idx, column=ec['col_idx']).value)
            balances[ec['date']] = val if val is not None else 0.0

        if uuid:
            INVESTOR_METADATA[uuid] = {'fee_pct': fee_pct, 'ib_pct': ib_pct}

        investors.append({
            'name': name_str,
            'uuid': uuid,
            'fee_pct': fee_pct,
            'ib_pct': ib_pct,
            'balances': balances,
        })

    print(f"  Investors: {len(investors)}")

    # Process events
    output_events = []
    positions = {}  # uuid -> float

    for em in events_meta:
        ev_date = em['date']
        gp = em['gross_pct']
        aum_after = em['aum_after']
        top_up = em['top_up']

        # Classify
        if is_zero(gp) or gp is None:
            if top_up and abs(top_up) > 0.01:
                ev_type = 'FLOW'
            else:
                print(f"  {fund_name}: Skipping zero-yield on {ev_date}")
                for inv in investors:
                    if inv['uuid']:
                        snap = inv['balances'].get(ev_date)
                        if snap is not None:
                            positions[inv['uuid']] = snap
                continue
        elif is_flow(gp):
            ev_type = 'FLOW'
        elif is_yield(gp):
            ev_type = 'YIELD'
        elif gp < 0:
            print(f"  {fund_name}: Skipping negative yield on {ev_date}")
            continue
        else:
            print(f"  {fund_name}: Unknown gross_pct={gp} on {ev_date}")
            continue

        if ev_type == 'FLOW':
            flow_txs = []
            for inv in investors:
                uuid = inv['uuid']
                if uuid is None:
                    continue

                # Skip investors who only receive yield (no real deposits/withdrawals)
                if uuid in YIELD_ONLY_INVESTORS:
                    snap = inv['balances'].get(ev_date)
                    if snap is not None:
                        positions[uuid] = snap
                    continue

                curr = inv['balances'].get(ev_date, 0.0)
                if curr is None:
                    curr = 0.0
                prev = positions.get(uuid, 0.0)

                if abs(curr - prev) < 0.0001:
                    positions[uuid] = curr
                    continue

                if prev < 0.0001 and curr > 0.01:
                    flow_txs.append({
                        'type': 'DEPOSIT',
                        'investor_name': inv['name'],
                        'investor_uuid': uuid,
                        'amount': fmt(curr),
                        'fee_pct': inv['fee_pct'],
                        'ib_pct': inv['ib_pct'],
                    })
                elif prev > 0.01 and curr < 0.0001:
                    flow_txs.append({
                        'type': 'WITHDRAWAL',
                        'investor_name': inv['name'],
                        'investor_uuid': uuid,
                        'amount': 'FULL',
                        'fee_pct': inv['fee_pct'],
                        'ib_pct': inv['ib_pct'],
                    })
                elif curr > prev + 0.0001:
                    flow_txs.append({
                        'type': 'DEPOSIT',
                        'investor_name': inv['name'],
                        'investor_uuid': uuid,
                        'amount': fmt(curr - prev),
                        'fee_pct': inv['fee_pct'],
                        'ib_pct': inv['ib_pct'],
                    })
                elif prev - curr > 0.0001:
                    flow_txs.append({
                        'type': 'WITHDRAWAL',
                        'investor_name': inv['name'],
                        'investor_uuid': uuid,
                        'amount': fmt(prev - curr),
                        'fee_pct': inv['fee_pct'],
                        'ib_pct': inv['ib_pct'],
                    })

                positions[uuid] = curr

            if flow_txs:
                output_events.append({
                    'fund': fund_name,
                    'fund_id': fund_id,
                    'event_type': 'FLOW',
                    'date': ev_date,
                    'aum_after': fmt(aum_after),
                    'comment': em['comment'],
                    'transactions': flow_txs,
                })
            else:
                print(f"  {fund_name}: FLOW on {ev_date} no transactions (comment: {em['comment'][:50]})")

        elif ev_type == 'YIELD':
            period_end = last_day(ev_date)

            # Check if this YIELD event also has deposits/withdrawals (top_up != 0)
            has_combined_flow = top_up is not None and abs(top_up) > 0.01

            if has_combined_flow:
                # Step 1: Apply yield mathematically to get pre-flow positions.
                # For investors with IB relationships, the Excel balance reflects total_deduction = fee + ib_pct.
                # Use fee_pct + ib_pct as the effective deduction to avoid spurious IB micro-adjustments.
                post_yield_positions = {}
                for inv in investors:
                    uuid = inv['uuid']
                    if uuid is None:
                        continue
                    prev = positions.get(uuid, 0.0)
                    if prev > 0.01:
                        fee = inv['fee_pct'] if inv['fee_pct'] is not None else 0.3
                        ib = inv['ib_pct'] if inv['ib_pct'] is not None else 0.0
                        total_deduction = fee + ib
                        net_rate = gp * (1.0 - total_deduction)
                        post_yield_positions[uuid] = prev * (1.0 + net_rate)
                    else:
                        post_yield_positions[uuid] = 0.0

                # Step 2: Find flow transactions by comparing post-yield vs actual snapshot
                flow_txs = []
                for inv in investors:
                    uuid = inv['uuid']
                    if uuid is None:
                        continue
                    if uuid in YIELD_ONLY_INVESTORS:
                        snap = inv['balances'].get(ev_date)
                        if snap is not None:
                            positions[uuid] = snap
                        continue

                    curr = inv['balances'].get(ev_date, 0.0)
                    if curr is None:
                        curr = 0.0
                    expected = post_yield_positions.get(uuid, 0.0)

                    # New deposit: investor had nothing, now has balance
                    if expected < 0.01 and curr > 0.01:
                        flow_txs.append({
                            'type': 'DEPOSIT',
                            'investor_name': inv['name'],
                            'investor_uuid': uuid,
                            'amount': fmt(curr),
                            'fee_pct': inv['fee_pct'],
                            'ib_pct': inv['ib_pct'],
                        })
                    # Full withdrawal
                    elif expected > 0.01 and curr < 0.01:
                        flow_txs.append({
                            'type': 'WITHDRAWAL',
                            'investor_name': inv['name'],
                            'investor_uuid': uuid,
                            'amount': 'FULL',
                            'fee_pct': inv['fee_pct'],
                            'ib_pct': inv['ib_pct'],
                        })
                    # Partial deposit or withdrawal
                    elif expected > 0.01 and curr > 0.01:
                        diff = curr - expected
                        abs_diff = abs(diff)
                        pct_diff = abs_diff / max(expected, 0.01)
                        if abs_diff > 1.0 or pct_diff > 0.005:
                            if diff > 0:
                                flow_txs.append({
                                    'type': 'DEPOSIT',
                                    'investor_name': inv['name'],
                                    'investor_uuid': uuid,
                                    'amount': fmt(diff),
                                    'fee_pct': inv['fee_pct'],
                                    'ib_pct': inv['ib_pct'],
                                })
                            else:
                                flow_txs.append({
                                    'type': 'WITHDRAWAL',
                                    'investor_name': inv['name'],
                                    'investor_uuid': uuid,
                                    'amount': fmt(abs_diff),
                                    'fee_pct': inv['fee_pct'],
                                    'ib_pct': inv['ib_pct'],
                                })

                    positions[uuid] = curr

                # Step 3: Emit YIELD event first, then FLOW event (if any)
                output_events.append({
                    'fund': fund_name,
                    'fund_id': fund_id,
                    'event_type': 'YIELD',
                    'date': ev_date,
                    'period_end': period_end,
                    'aum_after': fmt(aum_after),
                    'gross_pct': gp,
                    'comment': em['comment'],
                })

                if flow_txs:
                    output_events.append({
                        'fund': fund_name,
                        'fund_id': fund_id,
                        'event_type': 'FLOW',
                        'date': ev_date,
                        'aum_after': fmt(aum_after),
                        'comment': em['comment'],
                        'transactions': flow_txs,
                    })
                else:
                    print(f'  {fund_name}: YIELD+FLOW on {ev_date} but no flow txs detected (top_up={top_up})')

            else:
                # Pure yield event: update positions from snapshot
                for inv in investors:
                    if inv['uuid']:
                        snap = inv['balances'].get(ev_date)
                        if snap is not None:
                            positions[inv['uuid']] = snap

                output_events.append({
                    'fund': fund_name,
                    'fund_id': fund_id,
                    'event_type': 'YIELD',
                    'date': ev_date,
                    'period_end': period_end,
                    'aum_after': fmt(aum_after),
                    'gross_pct': gp,
                    'comment': em['comment'],
                })

    flow_count = sum(1 for e in output_events if e['event_type'] == 'FLOW')
    yield_count = sum(1 for e in output_events if e['event_type'] == 'YIELD')
    print(f"  {fund_name}: {len(output_events)} events ({flow_count} flow, {yield_count} yield)")
    return output_events


def parse_cf_final_balances(ws, fund_name):
    """Extract final balances from columns-first sheet."""
    last_col = None
    last_date = None
    for col_idx in range(4, ws.max_column + 1):
        cell = ws.cell(row=8, column=col_idx)
        if cell.value is not None and isinstance(cell.value, (datetime, date)):
            last_col = col_idx
            last_date = to_date_str(cell.value)

    if last_col is None:
        return {}

    result = {}
    for row_idx in range(9, ws.max_row + 1):
        name_raw = ws.cell(row=row_idx, column=1).value
        if name_raw is None:
            continue
        name_str = str(name_raw).strip()
        name_lower = normalize_name(name_str)
        if 'total' in name_lower:
            break
        if name_str == 'x' or name_lower in SKIP_INVESTORS:
            continue
        uuid = resolve_uuid(name_str)
        if uuid:
            bal = safe_float(ws.cell(row=row_idx, column=last_col).value)
            result[uuid] = {
                'name': name_str,
                'balance': bal if bal else 0.0,
                'as_of': last_date,
            }

    return result


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    all_events = []
    final_balances = {}

    ws_btc = wb['BTC Yield Fund']
    btc_events = parse_btc(ws_btc)
    all_events.extend(btc_events)
    final_balances['BTC'] = parse_btc_final_balances(ws_btc)

    ws_eth = wb['ETH Yield Fund']
    eth_events = parse_cf(ws_eth, 'ETH', FUND_IDS['ETH'])
    all_events.extend(eth_events)
    final_balances['ETH'] = parse_cf_final_balances(ws_eth, 'ETH')

    ws_usdt = wb['USDT Yield Fund']
    usdt_events = parse_cf(ws_usdt, 'USDT', FUND_IDS['USDT'])
    all_events.extend(usdt_events)
    final_balances['USDT'] = parse_cf_final_balances(ws_usdt, 'USDT')

    ws_sol = wb['SOL Yield Fund']
    sol_events = parse_cf(ws_sol, 'SOL', FUND_IDS['SOL'])
    all_events.extend(sol_events)
    final_balances['SOL'] = parse_cf_final_balances(ws_sol, 'SOL')

    ws_xrp = wb['XRP Yield Fund']
    xrp_events = parse_cf(ws_xrp, 'XRP', FUND_IDS['XRP'])
    all_events.extend(xrp_events)
    final_balances['XRP'] = parse_cf_final_balances(ws_xrp, 'XRP')

    all_events.sort(key=lambda e: e['date'])

    # Print BTC events for verification
    print("\n=== BTC Events ===")
    for ev in [e for e in all_events if e['fund'] == 'BTC']:
        if ev['event_type'] == 'FLOW':
            txs = ', '.join([f"{t['type']} {t['investor_name']} {t['amount']}" for t in ev['transactions']])
            print(f"  {ev['date']} FLOW: {txs}")
        else:
            print(f"  {ev['date']} YIELD gross={ev.get('gross_pct')} aum={ev['aum_after']}")

    # Summary
    print(f"\n=== Summary ===")
    print(f"Total events: {len(all_events)}")
    for fund in ['BTC', 'ETH', 'USDT', 'SOL', 'XRP']:
        fund_events = [e for e in all_events if e['fund'] == fund]
        flow_events = [e for e in fund_events if e['event_type'] == 'FLOW']
        yield_events = [e for e in fund_events if e['event_type'] == 'YIELD']
        total_txs = sum(len(e.get('transactions', [])) for e in flow_events)
        print(f"  {fund}: {len(fund_events)} total ({len(flow_events)} flow with {total_txs} txs, {len(yield_events)} yield)")

    output = {
        'admin_id': ADMIN_ID,
        'fund_ids': FUND_IDS,
        'investor_metadata': INVESTOR_METADATA,
        'events': all_events,
        'final_balances': final_balances,
    }

    with open(OUTPUT_PATH, 'w') as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\nWritten to: {OUTPUT_PATH}")


if __name__ == '__main__':
    main()
