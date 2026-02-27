"""
Analyze hardcoded manual adjustments in the Indigo Fees row across all fund sheets.

For each fund sheet, identifies columns where the Indigo Fees formula contains
hardcoded numeric constants or extra terms beyond the standard pattern:
  new_fees = prev_fees * gp + prev_fees + SUM(investor_prev * gp * fee_pct)

Outputs the exact amount, date, column number, and description for each adjustment.
"""

import re
import openpyxl
from decimal import Decimal, getcontext

getcontext().prec = 28

EXCEL_PATH = "scripts/seed-data/indigo-accounting.xlsx"


def analyze_btc(wb):
    """BTC Yield Fund: Rows-first layout. Indigo Fees = Row 2. Events start Col 12."""
    ws = wb["BTC Yield Fund"]
    adjustments = []

    # Standard pattern for BTC Indigo Fees (row 2):
    #   =PREV2*$C$GP*1+PREV2+(PREV_inv1*$C$GP*$J$inv1)+...
    # Look for hardcoded numeric literals added/subtracted

    for c in range(12, ws.max_column + 1):
        formula = ws.cell(row=2, column=c).value
        if formula is None:
            break
        if not isinstance(formula, str) or not formula.startswith("="):
            continue

        date = ws.cell(row=1, column=c).value

        # --- Col 25 (2025-04-16): +0.0498 ---
        # Transfer from BTC TAC Program. Indigo's share of TAC fund residual.
        if "+0.0498" in formula:
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": Decimal("0.0498"),
                "formula_snippet": "+0.0498",
                "description": (
                    "Transfer from BTC TAC Program closing. "
                    "Indigo's share of TAC fund residual yield (0.0498 BTC). "
                    "Other investors also received transfers: Kyle +2.101, "
                    "Matthias +4.8357, Danielle +5.0335."
                ),
            })

        # --- Col 51 (2025-12-23): Matthias withdrawal residual ---
        # (AX5*$C$69*(1-$J5)+AX5-4.9895)
        # Matthias (row 5) fully exits. His net balance after yield goes to Indigo,
        # minus his 4.9895 BTC withdrawal.
        if "-4.9895" in formula:
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": "DYNAMIC (Matthias_balance - 4.9895)",
                "formula_snippet": "(AX5*$C$69*(1-$J5)+AX5-4.9895)",
                "description": (
                    "Matthias Reiser full withdrawal. His entire net balance "
                    "after yield is absorbed by Indigo Fees, minus his 4.9895 BTC "
                    "withdrawal amount. The residual = accrued yield that stays "
                    "with Indigo."
                ),
            })

        # --- Col 54 (2026-01-05): Vivie & Liana withdrawal residual ---
        # (BA18*$C$74*(1-$J18)+BA18)-3.4221
        # Vivie & Liana (row 18, fee=0%) fully exit. Balance goes to Indigo minus withdrawal.
        if "-3.4221" in formula:
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": "DYNAMIC (Vivie_Liana_balance - 3.4221)",
                "formula_snippet": "(BA18*$C$74*(1-$J18)+BA18)-3.4221",
                "description": (
                    "Vivie & Liana full withdrawal. Their entire balance "
                    "after yield is absorbed by Indigo Fees (fee_pct=0%), "
                    "minus their 3.4221 BTC withdrawal amount. Residual = "
                    "accrued yield."
                ),
            })

    return adjustments


def analyze_eth(wb):
    """ETH Yield Fund: Columns-first layout. Indigo Fees = Row 9. Events start Col 4."""
    ws = wb["ETH Yield Fund"]
    adjustments = []

    for c in range(4, ws.max_column + 1):
        formula = ws.cell(row=9, column=c).value
        if formula is None:
            break
        if not isinstance(formula, str) or not formula.startswith("="):
            continue

        date = ws.cell(row=8, column=c).value

        # --- Col 7 (2025-07-11): +0.898 ---
        # Transfer from ETH TAC Program. Indigo's LP share = 0.898 ETH
        if "+0.898" in formula and "SUMIFS" not in formula:
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": Decimal("0.898"),
                "formula_snippet": "+0.898",
                "description": (
                    "Transfer from ETH TAC Program closing. Indigo's LP share "
                    "of the TAC fund = 0.898 ETH. Other transfers: Jose 62.6261, "
                    "Nathanael 26.6797, Blondish 119.7862."
                ),
            })

        # --- Col 23 (2025-11-03): +0.0359 ---
        # Tea-Fi hack refund. Indigo's share = 0.0359 ETH
        if "+0.0359" in formula:
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": Decimal("0.0359"),
                "formula_snippet": "+0.0359",
                "description": (
                    "Tea-Fi hack refund distribution. Indigo's share = 0.0359 ETH. "
                    "Total refund 8.404 ETH split: Jose 2.5064, Nathanael 1.0677, "
                    "Blondish 4.7940, Indigo 0.0359."
                ),
            })

        # --- Col 25 (2025-11-05): Paul Johnson exit residual ---
        # (X18*$Y$4*(1-$B18-$C18)+X18)-12.22
        if "-12.22" in formula:
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": "DYNAMIC (Paul_Johnson_balance - 12.22)",
                "formula_snippet": "(X18*$Y$4*(1-$B18-$C18)+X18)-12.22",
                "description": (
                    "Paul Johnson full withdrawal from ETH fund. His entire net "
                    "balance after yield (fee=13.5%, IB=1.5%) is absorbed by Indigo "
                    "Fees, minus his 12.22 ETH withdrawal amount."
                ),
            })

        # --- Col 35 (2026-01-02): Sam Johnson exit residual ---
        # -213.73+(AH21*$AI$4*(1-$B21-$C21)+AH21)
        if "-213.73" in formula:
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": "DYNAMIC (Sam_Johnson_balance - 213.73)",
                "formula_snippet": "-213.73+(AH21*$AI$4*(1-$B21-$C21)+AH21)",
                "description": (
                    "Sam Johnson full withdrawal from ETH fund. His entire net "
                    "balance after yield (fee=16%, IB=4%) is absorbed by Indigo "
                    "Fees, minus his 213.73 ETH withdrawal amount."
                ),
            })

        # --- Cols 36-37: SUMIFS for Indigo Fees deposits ---
        # These are NOT hardcoded; they pull from the Investments sheet dynamically.
        # Including them for completeness but flagging as non-hardcoded.
        if "SUMIFS" in str(formula) and c in (36, 37):
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": "DYNAMIC (from Investments sheet)",
                "formula_snippet": "SUMIFS(Investments!...)",
                "description": (
                    "Dynamic deposit/withdrawal lookup from Investments sheet for "
                    "Indigo Fees account. NOT a hardcoded adjustment - included "
                    "for completeness."
                ),
            })

    return adjustments


def analyze_usdt(wb):
    """USDT Yield Fund: Columns-first layout. Indigo Fees = Row 9."""
    ws = wb["USDT Yield Fund"]
    adjustments = []

    for c in range(4, ws.max_column + 1):
        formula = ws.cell(row=9, column=c).value
        if formula is None:
            break
        if not isinstance(formula, str) or not formula.startswith("="):
            continue

        date = ws.cell(row=8, column=c).value

        # --- Col 22 (2025-11-03): +26.13 ---
        # INDIGO LP withdrawal delta. LP account worth 113,867.78, withdrew 113,841.65.
        # Indigo keeps the 26.13 USDT delta.
        if "+26.13" in formula and "SUMIFS" not in formula:
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": Decimal("26.13"),
                "formula_snippet": "+26.13",
                "description": (
                    "INDIGO LP account withdrawal delta. LP account worth "
                    "113,867.78 USDT, withdrew 113,841.65 USDT. Indigo keeps "
                    "the 26.13 USDT residual difference."
                ),
            })

        # --- Col 34 (2026-01-08): Daniele Francilia + INDIGO Ventures exit ---
        # +AG12*AH$4*(1-$B12)+AG12-114867.59+AG20*AH$4*(1-$B20)+AG20-5115.04
        if "-114867.59" in formula:
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": "DYNAMIC (Daniele_balance - 114867.59 + INDIGO_Ventures_balance - 5115.04)",
                "formula_snippet": "AG12*AH$4*(1-$B12)+AG12-114867.59+AG20*AH$4*(1-$B20)+AG20-5115.04",
                "description": (
                    "Two full withdrawals on same date: "
                    "(1) Daniele Francilia (fee=10%) exits with 114,867.59 USDT "
                    "withdrawal, residual accrued yield absorbed by Indigo; "
                    "(2) INDIGO Ventures (fee=0%) exits with 5,115.04 USDT "
                    "withdrawal, residual absorbed by Indigo."
                ),
            })

        # SUMIFS-based deposits (dynamic, not hardcoded)
        if "SUMIFS" in str(formula) and "-114867.59" not in formula:
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": "DYNAMIC (from Investments sheet)",
                "formula_snippet": "SUMIFS(Investments!...)",
                "description": (
                    "Dynamic deposit/withdrawal lookup from Investments sheet for "
                    "Indigo Fees account. NOT a hardcoded adjustment."
                ),
            })

    return adjustments


def analyze_sol(wb):
    """SOL Yield Fund: Columns-first layout. Indigo Fees = Row 9."""
    ws = wb["SOL Yield Fund"]
    adjustments = []

    for c in range(4, ws.max_column + 1):
        formula = ws.cell(row=9, column=c).value
        if formula is None:
            break
        if not isinstance(formula, str) or not formula.startswith("="):
            continue

        date = ws.cell(row=8, column=c).value

        # --- Col 7 (2025-10-03): Paul Johnson exit residual ---
        # (F11*$G$4*(1-$B11-$C11)+F11)-236.02
        if "-236.02" in formula:
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": "DYNAMIC (Paul_Johnson_balance - 236.02)",
                "formula_snippet": "(F11*$G$4*(1-$B11-$C11)+F11)-236.02",
                "description": (
                    "Paul Johnson full withdrawal from SOL fund. His entire net "
                    "balance after yield (fee=10%) is absorbed by Indigo Fees, "
                    "minus his 236.02 SOL withdrawal amount."
                ),
            })

        # --- Col 13 (2025-12-04): INDIGO DIGITAL ASSET FUND LP exit ---
        # (L10*$M$4+L10-1285.66)
        if "-1285.66" in formula:
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": "DYNAMIC (INDIGO_LP_balance - 1285.66)",
                "formula_snippet": "(L10*$M$4+L10-1285.66)",
                "description": (
                    "INDIGO DIGITAL ASSET FUND LP full withdrawal from SOL fund. "
                    "LP balance after yield (fee=0%) absorbed by Indigo Fees, "
                    "minus 1,285.66 SOL withdrawal amount. LP row zeroed out after."
                ),
            })

        # --- Col 17 (2026-01-02): Sam Johnson exit residual ---
        # (P14*$Q$4*(1-$B14-$C14)+P14)-4873.15
        if "-4873.15" in formula:
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": "DYNAMIC (Sam_Johnson_balance - 4873.15)",
                "formula_snippet": "(P14*$Q$4*(1-$B14-$C14)+P14)-4873.15",
                "description": (
                    "Sam Johnson full withdrawal from SOL fund. His entire net "
                    "balance after yield (fee=16%, IB=4%) is absorbed by Indigo "
                    "Fees, minus his 4,873.15 SOL withdrawal amount."
                ),
            })

    return adjustments


def analyze_xrp(wb):
    """XRP Yield Fund: Columns-first layout. Indigo Fees = Row 9."""
    ws = wb["XRP Yield Fund"]
    adjustments = []

    for c in range(4, ws.max_column + 1):
        formula = ws.cell(row=9, column=c).value
        if formula is None:
            break
        if not isinstance(formula, str) or not formula.startswith("="):
            continue

        date = ws.cell(row=8, column=c).value

        # --- Col 11 (2026-01-05): (792-475.58)*80% = 253.136 ---
        # Special yield split: After Sam Johnson fully exits (col 10), the
        # remaining AUM grows from 475.58 to 792 (net of deposits). Indigo gets
        # 80% of the growth, Ryan Van Der Wall gets 20%.
        if "(792-475.58)*80%" in formula:
            amount = (Decimal("792") - Decimal("475.58")) * Decimal("0.80")
            adjustments.append({
                "col": c,
                "date": str(date),
                "amount": amount,
                "formula_snippet": "(792-475.58)*80%",
                "description": (
                    "Special yield split after Sam Johnson exit. AUM grows from "
                    "475.58 to 792 XRP. Indigo gets 80% of the growth "
                    f"= {amount} XRP. Ryan Van Der Wall gets 20%. "
                    "Comment: 'INDIGO Fees gets 80% of the yield'."
                ),
            })

    return adjustments


def compute_btc_hardcoded_net(wb):
    """Compute the net hardcoded amount for BTC by evaluating the formulas with data."""
    ws = wb["BTC Yield Fund"]
    # Load the workbook with data_only to get computed values
    wb_data = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws_data = wb_data["BTC Yield Fund"]

    results = {}

    # Col 25: +0.0498 (simple hardcoded constant)
    results["col_25"] = {"amount": Decimal("0.0498"), "date": "2025-04-16"}

    # Col 51: Matthias residual = (balance_after_yield - 4.9895)
    # balance = AX5*$C$69*(1-$J5)+AX5
    # We need AX5 = col 50, row 5 value; C69 = GP; J5 = fee_pct = 0.1
    ax5 = ws_data.cell(row=5, column=50).value  # Matthias at col 50
    c69 = ws_data.cell(row=69, column=3).value   # GP for that period
    j5 = Decimal("0.1")  # fee_pct
    if ax5 is not None and c69 is not None:
        ax5 = Decimal(str(ax5))
        c69 = Decimal(str(c69))
        balance = ax5 * c69 * (1 - j5) + ax5
        residual = balance - Decimal("4.9895")
        results["col_51"] = {
            "amount": residual,
            "date": "2025-12-23",
            "balance": balance,
            "withdrawal": Decimal("4.9895"),
        }
    else:
        results["col_51"] = {"amount": "COULD NOT COMPUTE (data_only values missing)", "date": "2025-12-23"}

    # Col 54: Vivie & Liana residual = (balance_after_yield - 3.4221)
    # balance = BA18*$C$74*(1-$J18)+BA18; J18 = 0.0
    ba18 = ws_data.cell(row=18, column=53).value  # Vivie at col 53
    c74 = ws_data.cell(row=74, column=3).value     # GP
    j18 = Decimal("0.0")
    if ba18 is not None and c74 is not None:
        ba18 = Decimal(str(ba18))
        c74 = Decimal(str(c74))
        balance = ba18 * c74 * (1 - j18) + ba18
        residual = balance - Decimal("3.4221")
        results["col_54"] = {
            "amount": residual,
            "date": "2026-01-05",
            "balance": balance,
            "withdrawal": Decimal("3.4221"),
        }
    else:
        results["col_54"] = {"amount": "COULD NOT COMPUTE", "date": "2026-01-05"}

    wb_data.close()
    return results


def compute_all_residuals(wb):
    """Compute the exact residual amounts for all dynamic adjustments."""
    wb_data = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    results = {}

    # --- BTC ---
    ws = wb_data["BTC Yield Fund"]
    # Col 51: Matthias (row 5) - prev col is 50
    ax5 = ws.cell(row=5, column=50).value
    c69 = ws.cell(row=69, column=3).value
    if ax5 and c69:
        ax5, c69 = Decimal(str(ax5)), Decimal(str(c69))
        bal = ax5 * c69 * Decimal("0.9") + ax5
        results["BTC_col51_Matthias"] = {"balance": bal, "withdrawal": Decimal("4.9895"), "residual": bal - Decimal("4.9895")}

    # Col 54: Vivie & Liana (row 18) - prev col is 53
    ba18 = ws.cell(row=18, column=53).value
    c74 = ws.cell(row=74, column=3).value
    if ba18 and c74:
        ba18, c74 = Decimal(str(ba18)), Decimal(str(c74))
        bal = ba18 * c74 * Decimal("1.0") + ba18  # fee=0
        results["BTC_col54_Vivie"] = {"balance": bal, "withdrawal": Decimal("3.4221"), "residual": bal - Decimal("3.4221")}

    # --- ETH ---
    ws = wb_data["ETH Yield Fund"]
    # Col 25: Paul Johnson (row 18) - prev col is 24 (col X)
    x18 = ws.cell(row=18, column=24).value
    y4 = ws.cell(row=4, column=25).value
    if x18 and y4:
        x18, y4 = Decimal(str(x18)), Decimal(str(y4))
        bal = x18 * y4 * (1 - Decimal("0.135") - Decimal("0.015")) + x18
        results["ETH_col25_Paul"] = {"balance": bal, "withdrawal": Decimal("12.22"), "residual": bal - Decimal("12.22")}

    # Col 35: Sam Johnson (row 21) - prev col is 34
    ah21 = ws.cell(row=21, column=34).value
    ai4 = ws.cell(row=4, column=35).value
    if ah21 and ai4:
        ah21, ai4 = Decimal(str(ah21)), Decimal(str(ai4))
        bal = ah21 * ai4 * (1 - Decimal("0.16") - Decimal("0.04")) + ah21
        results["ETH_col35_Sam"] = {"balance": bal, "withdrawal": Decimal("213.73"), "residual": bal - Decimal("213.73")}

    # --- SOL ---
    ws = wb_data["SOL Yield Fund"]
    # Col 7: Paul Johnson (row 11) - prev col is 6
    f11 = ws.cell(row=11, column=6).value
    g4 = ws.cell(row=4, column=7).value
    if f11 and g4:
        f11, g4 = Decimal(str(f11)), Decimal(str(g4))
        # B11=0.1, C11=None->0
        bal = f11 * g4 * (1 - Decimal("0.1")) + f11
        results["SOL_col7_Paul"] = {"balance": bal, "withdrawal": Decimal("236.02"), "residual": bal - Decimal("236.02")}

    # Col 13: INDIGO LP (row 10) - prev col is 12
    l10 = ws.cell(row=10, column=12).value
    m4 = ws.cell(row=4, column=13).value
    if l10 and m4:
        l10, m4 = Decimal(str(l10)), Decimal(str(m4))
        bal = l10 * m4 + l10  # fee=0
        results["SOL_col13_LP"] = {"balance": bal, "withdrawal": Decimal("1285.66"), "residual": bal - Decimal("1285.66")}

    # Col 17: Sam Johnson (row 14) - prev col is 16
    p14 = ws.cell(row=14, column=16).value
    q4 = ws.cell(row=4, column=17).value
    if p14 and q4:
        p14, q4 = Decimal(str(p14)), Decimal(str(q4))
        # B14=0.16, C14=0.04
        bal = p14 * q4 * (1 - Decimal("0.16") - Decimal("0.04")) + p14
        results["SOL_col17_Sam"] = {"balance": bal, "withdrawal": Decimal("4873.15"), "residual": bal - Decimal("4873.15")}

    # --- USDT ---
    ws = wb_data["USDT Yield Fund"]
    # Col 34: Daniele (row 12) + INDIGO Ventures (row 20)
    ag12 = ws.cell(row=12, column=33).value  # prev col
    ah4 = ws.cell(row=4, column=34).value
    ag20 = ws.cell(row=20, column=33).value
    if ag12 and ah4 and ag20:
        ag12, ah4, ag20 = Decimal(str(ag12)), Decimal(str(ah4)), Decimal(str(ag20))
        bal_daniele = ag12 * ah4 * (1 - Decimal("0.1")) + ag12
        bal_ventures = ag20 * ah4 * Decimal("1.0") + ag20  # fee=0
        res_daniele = bal_daniele - Decimal("114867.59")
        res_ventures = bal_ventures - Decimal("5115.04")
        results["USDT_col34_Daniele"] = {"balance": bal_daniele, "withdrawal": Decimal("114867.59"), "residual": res_daniele}
        results["USDT_col34_Ventures"] = {"balance": bal_ventures, "withdrawal": Decimal("5115.04"), "residual": res_ventures}

    wb_data.close()
    return results


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    print("=" * 100)
    print("HARDCODED MANUAL ADJUSTMENTS IN INDIGO FEES ROW")
    print("=" * 100)

    fund_analyses = {
        "BTC Yield Fund": analyze_btc(wb),
        "ETH Yield Fund": analyze_eth(wb),
        "USDT Yield Fund": analyze_usdt(wb),
        "SOL Yield Fund": analyze_sol(wb),
        "XRP Yield Fund": analyze_xrp(wb),
    }

    # Compute exact residual amounts
    residuals = compute_all_residuals(wb)

    for fund_name, adjustments in fund_analyses.items():
        print(f"\n{'=' * 80}")
        print(f"  {fund_name}")
        print(f"{'=' * 80}")

        hardcoded = [a for a in adjustments if "NOT a hardcoded" not in a["description"]]
        dynamic_deposits = [a for a in adjustments if "NOT a hardcoded" in a["description"]]

        if not hardcoded:
            print("  No hardcoded adjustments found.")
            continue

        for i, adj in enumerate(hardcoded, 1):
            print(f"\n  [{i}] Col {adj['col']} | Date: {adj['date']}")
            print(f"      Formula snippet: {adj['formula_snippet']}")
            print(f"      Amount: {adj['amount']}")
            print(f"      Description: {adj['description']}")

        if dynamic_deposits:
            print(f"\n  --- Dynamic (non-hardcoded) deposits via SUMIFS ---")
            for adj in dynamic_deposits:
                print(f"  Col {adj['col']} ({adj['date']}): {adj['description']}")

    # Summary of exact computed residuals
    print(f"\n\n{'=' * 100}")
    print("COMPUTED RESIDUAL AMOUNTS (from data_only workbook)")
    print("=" * 100)

    for key, data in sorted(residuals.items()):
        print(f"\n  {key}:")
        print(f"    Balance after yield: {data['balance']}")
        print(f"    Withdrawal amount:   {data['withdrawal']}")
        print(f"    Residual to Indigo:  {data['residual']}")

    # Compute total hardcoded amounts per fund
    print(f"\n\n{'=' * 100}")
    print("TOTAL HARDCODED AMOUNTS PER FUND (sum of all adjustments to Indigo Fees)")
    print("=" * 100)

    # BTC
    btc_total = Decimal("0.0498")
    if "BTC_col51_Matthias" in residuals:
        btc_total += residuals["BTC_col51_Matthias"]["residual"]
    if "BTC_col54_Vivie" in residuals:
        btc_total += residuals["BTC_col54_Vivie"]["residual"]
    print(f"\n  BTC: {btc_total}")
    print(f"    Known shortfall: -0.0000309641")
    print(f"    Difference:      {btc_total - Decimal('0.0000309641')} (this is what the engine should produce without the adjustments)")

    # ETH
    eth_total = Decimal("0.898") + Decimal("0.0359")
    if "ETH_col25_Paul" in residuals:
        eth_total += residuals["ETH_col25_Paul"]["residual"]
    if "ETH_col35_Sam" in residuals:
        eth_total += residuals["ETH_col35_Sam"]["residual"]
    print(f"\n  ETH: {eth_total}")
    print(f"    Known shortfall: -0.0336523133")

    # USDT
    usdt_total = Decimal("26.13")
    if "USDT_col34_Daniele" in residuals:
        usdt_total += residuals["USDT_col34_Daniele"]["residual"]
    if "USDT_col34_Ventures" in residuals:
        usdt_total += residuals["USDT_col34_Ventures"]["residual"]
    print(f"\n  USDT: {usdt_total}")
    print(f"    Known shortfall: -26.6515943615")

    # SOL
    sol_total = Decimal("0")
    if "SOL_col7_Paul" in residuals:
        sol_total += residuals["SOL_col7_Paul"]["residual"]
    if "SOL_col13_LP" in residuals:
        sol_total += residuals["SOL_col13_LP"]["residual"]
    if "SOL_col17_Sam" in residuals:
        sol_total += residuals["SOL_col17_Sam"]["residual"]
    print(f"\n  SOL: {sol_total}")
    print(f"    Known shortfall: -0.0026488551")

    # XRP
    xrp_total = (Decimal("792") - Decimal("475.58")) * Decimal("0.80")
    print(f"\n  XRP: {xrp_total}")
    print(f"    Known shortfall: -0.0005867418")

    wb.close()


if __name__ == "__main__":
    main()
